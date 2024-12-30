import numpy as np
import json
from collections import defaultdict
import argparse
import re
import math
import os
import ast
import logging
import openpyxl

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

TYPE="high"
MODEL="qwen2vl"
PRED_FILE_PATH=f"results/{MODEL}/predictions.jsonl"
def get_in_domain_ids(jsonl_path):
    in_domain_ids = []
    with open(jsonl_path, 'r') as f:
        for line in f:
            data = json.loads(line)
            in_domain_ids.append(data["id"])
    return in_domain_ids


def extract_json(s):
    match = re.search(r'\{.*?\}', s)
    if match:
        json_str = match.group(0)
        try:
            return json.loads(json_str)
        except json.JSONDecodeError:
            return json_str  # 返回字符串形式
    return None


def extract_dict_from_string(text):
    start_index = text.find("Accessibility tree: ")
    if start_index == -1:
        return None  # 没有找到 "Accessibility tree: " 返回 None
    
    # 找到字典的起始位置
    dict_start = start_index + len("Accessibility tree: ")
    
    # 提取出字典部分的字符串
    dict_str = text[dict_start:]
    
    # 找到第一个 '{' 和匹配的 '}'
    open_brace_index = dict_str.find('{')
    if open_brace_index == -1:
        return None  # 没有找到 '{' 返回 None
    
    stack = []
    for i, char in enumerate(dict_str[open_brace_index:], start=open_brace_index):
        if char == '{':
            stack.append(char)
        elif char == '}':
            stack.pop()
            if not stack:
                dict_end = i + 1
                break
    else:
        return None  # 没有找到匹配的 '}' 返回 None
    
    # 提取出字典字符串
    dict_substr = dict_str[open_brace_index:dict_end]
    
    return dict_substr

def get_key_by_position(text, x, y):
    position = f"({x}, {y})"
    extracted_dict = json.loads(extract_dict_from_string(text))
    if extracted_dict is None:
        return None  # 如果没有找到字典或解析失败返回 None
    
    for key in extracted_dict:
        if extracted_dict[key] == position:
            return key
    
    return None  # 没有找到匹配的位置返回 None

def calculate_f1_score(predicted_str, ground_truth_str):
    predicted_tokens = set(predicted_str.lower().split())
    ground_truth_tokens = set(ground_truth_str.lower().split())
    
    common_tokens = predicted_tokens.intersection(ground_truth_tokens)
    if len(predicted_tokens) == 0:
        precision = 0
    else:
        precision = len(common_tokens) / len(predicted_tokens)
    if len(ground_truth_tokens) == 0:
        recall = 0
    else:
        recall = len(common_tokens) / len(ground_truth_tokens)
    
    if precision + recall == 0:
        f1_score = 0
    else:
        f1_score = 2 * (precision * recall) / (precision + recall)
    return f1_score


def evaluate(args):
    prediction_file_path = args.prediction_file_path
    prediction = []
    with open(prediction_file_path) as file:
        for line in file:
            prediction.append(json.loads(line))

    ground_truth = []
    with open("eval_json_files/android_control_test_data.jsonl") as file:
        for line in file:
            ground_truth.append(json.loads(line))

    with open(f"eval_json_files/android_control_test_subsplits.json","r") as file:
        test_subsplits = json.load(file)
    print(test_subsplits.keys())
    print(len(ground_truth))


    # ======================================================================== #
    #                          Results on Low-level
    # ======================================================================== #
    mis_click_wait_num = 0
    step_acc_res_dict = defaultdict(int)
    sample_number_dict = defaultdict(int)
    for pred, gt in zip(prediction, ground_truth):
        gt_action = json.loads(gt["conversations"][1]["value"].split("actions:\n")[1])
        episode_id = int(pred["image_id"].split("/")[-1].split("_")[1]) # parse out the episode index
        subsplit_type = next((category for category, ids in test_subsplits.items() if episode_id in ids), None)       
        gt_action_type = gt_action["action_type"] 
        sample_number_dict[subsplit_type+"_LL"] += 1
        sample_number_dict["full_LL"] += 1
        sample_number_dict["Type_LL"] += 1
        sample_number_dict[gt_action_type+"_LL"] += 1        

        if len(pred["pred"].split("action: "))==2:
            try:
                pred_action = json.loads(pred["pred"].split("action: ")[1])
                if len(pred_action) == 0:
                    continue
            except json.JSONDecodeError as e:
                continue
        else:
            pred_action = extract_json(pred["pred"])
            if pred_action is None:
                continue
        try:
            pred_action_type = pred_action["action_type"]
        except Exception as e:
            continue        

        # calculate step acc based on types
        if gt_action_type==pred_action_type or (gt_action_type == "type" and pred_action_type == "input_text"):
            step_acc_res_dict["Type_LL"] += 1
            step_acc_res_dict[gt_action_type+"_type_match_LL"] += 1
            if gt_action_type in ["click","long_press"]:  # evaluate click type
                try:
                    pred_x, pred_y = int(pred_action["x"]), int(pred_action["y"])
                except:
                    pred_x, pred_y = -100, -100
                gt_x, gt_y = int(gt_action["x"]), int(gt_action["y"])

                if math.sqrt((pred_x - gt_x)**2 + (pred_y - gt_y)**2) <= math.sqrt((1080*0.14)**2 + (2400*0.14)**2):  # set 14 % of screen size as the ratio
                    step_acc_res_dict[subsplit_type+"_LL"] += 1
                    step_acc_res_dict["full_LL"] += 1
                    step_acc_res_dict[gt_action_type+"_all_match_LL"] += 1


            elif gt_action_type == "type" and pred_action_type in ["input_text", "type"]:
                if gt_action["text"]==pred_action["text"] or calculate_f1_score(pred_action["text"], gt_action["text"])>0.5:
                    step_acc_res_dict[subsplit_type+"_LL"] += 1
                    step_acc_res_dict["full_LL"] += 1
                    step_acc_res_dict[gt_action_type+"_all_match_LL"] += 1

            elif gt_action_type == "scroll":
                if "Scroll up" in pred["prompt"] and pred_action["direction"] == "up":
                    step_acc_res_dict[subsplit_type+"_LL"] += 1
                    step_acc_res_dict["full_LL"] += 1
                    step_acc_res_dict[gt_action_type+"_all_match_LL"] += 1
                elif "Scroll down" in pred["prompt"] and pred_action["direction"] == "down":
                    step_acc_res_dict[subsplit_type+"_LL"] += 1
                    step_acc_res_dict["full_LL"] += 1
                    step_acc_res_dict[gt_action_type+"_all_match_LL"] += 1
                elif pred_action==gt_action:
                    step_acc_res_dict[subsplit_type+"_LL"] += 1
                    step_acc_res_dict["full_LL"] += 1
                    step_acc_res_dict[gt_action_type+"_all_match_LL"] += 1
                  
            elif gt_action==pred_action:  # evaluate other types
                step_acc_res_dict[subsplit_type+"_LL"] += 1
                step_acc_res_dict["full_LL"] += 1
                step_acc_res_dict[gt_action_type+"_all_match_LL"] += 1


    # Print the low-level results
    logger.info("="*30 + " Step Acc " + "="*30)
    logger.info("Full-LL: %f\n" % (step_acc_res_dict["full_LL"] / sample_number_dict["full_LL"]))
    logger.info("Type-LL: %f\n" % (step_acc_res_dict["Type_LL"] / sample_number_dict["Type_LL"]))
    # 保存结果到excel中
    SR_acc = round((step_acc_res_dict["full_LL"] / sample_number_dict["full_LL"]) * 100, 2)
    Type_acc = round((step_acc_res_dict["Type_LL"] / sample_number_dict["Type_LL"]) * 100, 2)
    # 打开 Excel 文件
    file_path = "android_control_eval.xlsx"
    wb = openpyxl.load_workbook(file_path)

    # 选择工作表
    sheet = wb.active

    # 找到下一个空行
    next_row = sheet.max_row + 1

    # 写入数据
    sheet.cell(row=next_row, column=1, value=MODEL)
    sheet.cell(row=next_row, column=2, value=SR_acc)
    sheet.cell(row=next_row, column=3, value=Type_acc)

    # 保存文件
    wb.save(file_path)    

    logger.info("IDD-LL: %f\n" % (step_acc_res_dict["IDD_LL"] / sample_number_dict["IDD_LL"]))
    logger.info("app_unseen-LL: %f\n" % (step_acc_res_dict["app_unseen_LL"] / sample_number_dict["app_unseen_LL"]))
    logger.info("task_unseen-LL: %f\n" % (step_acc_res_dict["task_unseen_LL"] / sample_number_dict["task_unseen_LL"]))
    logger.info("category_unseen-LL: %f\n" % (step_acc_res_dict["category_unseen_LL"] / sample_number_dict["category_unseen_LL"]))
    logger.info("="*30 + " Detailed Acc of Each Type " + "="*30)
    for action_type in sample_number_dict:
        action_type = action_type.split("_LL")[0]
        if action_type not in ["full","Type","IDD","app_unseen","task_unseen","category_unseen"]:
            logger.info(f"{action_type}_all_match-LL: %f\n" % (step_acc_res_dict[f"{action_type}_all_match_LL"] / sample_number_dict[f"{action_type}_LL"]))
            logger.info(f"{action_type}_type_match-LL: %f\n" % (step_acc_res_dict[f"{action_type}_type_match_LL"] / sample_number_dict[f"{action_type}_LL"]))



if __name__ == '__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument('--prediction_file_path', type=str, default=PRED_FILE_PATH)
    parser.add_argument('--datasets', type=str, default='')
    parser.add_argument('--output_path', type=str, default='results/score/')
    parser.add_argument('--eval_HH', action='store_true')
    parser.add_argument('--seed', type=int, default=0)
    args = parser.parse_args()

    if not os.path.exists(args.output_path):
        os.makedirs(args.output_path)

    file_handler = logging.FileHandler(args.output_path + f"ac_{TYPE}_score_{MODEL}.log", mode="w")
    file_handler.setLevel(logging.INFO)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    evaluate(args)